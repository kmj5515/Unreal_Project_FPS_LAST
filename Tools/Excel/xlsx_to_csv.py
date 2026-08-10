#!/usr/bin/env python3
"""Excel 워크시트를 Unreal 임포트용 CSV로 변환한다."""

from __future__ import annotations

import argparse
import csv
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import load_workbook


GENERATED_NOTICE = "# 자동 생성됨 - 직접 수정하지 말고 원본 xlsx를 편집하세요."
LOCALIZE_COLUMNS = ("Key", "SourceString", "Comment")


def cell_to_text(value: object) -> str:
    """셀 형식을 CSV에서 손실 없이 재현할 수 있는 문자열로 정규화한다."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "True" if value else "False"
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, (date, time)):
        return value.isoformat()
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return format(value, ".15g")
    if isinstance(value, Decimal):
        return format(value, "f")
    return str(value)


def _trim_trailing_empty(values: Sequence[str]) -> list[str]:
    result = list(values)
    while result and result[-1] == "":
        result.pop()
    return result


def read_sheet_rows(workbook_path: Path, sheet_name: str) -> list[list[str]]:
    """수식은 엑셀이 마지막으로 저장한 캐시값을 사용해 읽는다."""
    try:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    except PermissionError as exc:
        raise RuntimeError(f"엑셀 파일이 열려 있거나 잠겨 있습니다: {workbook_path}") from exc
    except OSError as exc:
        raise RuntimeError(f"엑셀 파일을 읽을 수 없습니다: {workbook_path}: {exc}") from exc

    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"워크시트를 찾을 수 없습니다: {workbook_path}::{sheet_name}")

        worksheet = workbook[sheet_name]
        rows = [
            _trim_trailing_empty([cell_to_text(value) for value in row])
            for row in worksheet.iter_rows(values_only=True)
        ]
    finally:
        workbook.close()

    while rows and not any(rows[-1]):
        rows.pop()
    if not rows:
        raise ValueError(f"워크시트가 비어 있습니다: {workbook_path}::{sheet_name}")
    return rows


def prepare_import_rows(rows: Sequence[Sequence[str]], *, localize: bool = False) -> list[list[str]]:
    """주석 행과 엑셀 전용 컬럼을 제거해 임포트 데이터 계약을 만든다."""
    header = list(rows[0])
    if not header:
        raise ValueError("헤더 행이 비어 있습니다.")

    if localize:
        if tuple(header[: len(LOCALIZE_COLUMNS)]) != LOCALIZE_COLUMNS:
            raise ValueError(
                "Localize 시트의 앞 3개 컬럼은 Key,SourceString,Comment 이어야 합니다. "
                f"현재 헤더={header}"
            )
        selected_indexes = list(range(len(LOCALIZE_COLUMNS)))
    else:
        selected_indexes = [index for index, name in enumerate(header) if not name.startswith("#")]

    if not selected_indexes:
        raise ValueError("임포트할 컬럼이 없습니다.")

    result = [[header[index] for index in selected_indexes]]
    for source_row in rows[1:]:
        first_cell = source_row[0] if source_row else ""
        if first_cell.startswith("#"):
            continue
        result.append([
            source_row[index] if index < len(source_row) else ""
            for index in selected_indexes
        ])
    return result


def write_csv(rows: Iterable[Sequence[str]], output_path: Path, *, include_notice: bool = True) -> None:
    """Unreal과 Excel 양쪽에서 안정적인 UTF-8 BOM CSV를 기록한다."""
    materialized = [list(row) for row in rows]
    if not materialized:
        raise ValueError("기록할 CSV 행이 없습니다.")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        with output_path.open("w", encoding="utf-8-sig", newline="") as stream:
            writer = csv.writer(stream, lineterminator="\n")
            writer.writerow(materialized[0])
            if include_notice:
                writer.writerow([GENERATED_NOTICE, *([""] * (len(materialized[0]) - 1))])
            writer.writerows(materialized[1:])
    except PermissionError as exc:
        raise RuntimeError(f"CSV 파일이 열려 있거나 잠겨 있습니다: {output_path}") from exc
    except OSError as exc:
        raise RuntimeError(f"CSV 파일을 기록할 수 없습니다: {output_path}: {exc}") from exc


def convert_sheet(
    workbook_path: Path,
    sheet_name: str,
    output_path: Path,
    *,
    localize: bool = False,
    include_notice: bool = True,
) -> list[list[str]]:
    rows = prepare_import_rows(read_sheet_rows(workbook_path, sheet_name), localize=localize)
    write_csv(rows, output_path, include_notice=include_notice)
    return rows


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path)
    parser.add_argument("sheet")
    parser.add_argument("output", type=Path)
    parser.add_argument("--localize", action="store_true", help="앞 3개 StringTable 컬럼만 출력합니다.")
    parser.add_argument("--no-notice", action="store_true", help="자동 생성 안내 행을 생략합니다.")
    return parser


def main() -> int:
    args = build_parser().parse_args()
    convert_sheet(
        args.workbook.resolve(),
        args.sheet,
        args.output.resolve(),
        localize=args.localize,
        include_notice=not args.no_notice,
    )
    print(f"변환 완료: {args.workbook}::{args.sheet} -> {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
