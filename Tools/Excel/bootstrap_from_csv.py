#!/usr/bin/env python3
"""기존 CSV 원본으로 Core/Hub/Battle 워크북을 최초 생성한다."""

from __future__ import annotations

import argparse
import csv
import io
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


WORKBOOK_SHEETS: dict[str, tuple[str, ...]] = {
    "Core.xlsx": (
        "DT_CharacterMaster",
        "DT_CharacterSkills",
        "DT_SkillBalance",
        "DT_WeaponBalance",
        "DT_StatusEffectData",
        "DT_LoadingTips",
    ),
    "Hub.xlsx": (
        "DT_QuestData",
        "DT_NPCData",
        "DT_DialogueData",
        "DT_ItemData",
        "DT_ShopData",
        "DT_RadioTransmission",
    ),
    "Battle.xlsx": ("DT_RoomEncounters",),
}

README_ROWS = (
    ("엑셀 데이터 편집 규칙", ""),
    ("데이터 시트", "1행은 컬럼명, 1열은 RowName입니다."),
    ("주석 행", "첫 셀이 #으로 시작하면 CSV 변환에서 제외됩니다."),
    ("계산 컬럼", "컬럼명이 #으로 시작하면 CSV 변환에서 제외됩니다."),
    ("제외 시트", "_Ref_, _Calc_, _Draft_, _Note, _README 접두사 시트는 임포트하지 않습니다."),
    ("새 데이터 시트", "임포트하려면 DataTable Import Registry에 등록하고, 제외하려면 이름 앞에 _를 붙입니다."),
    ("수식 캐시", "수식 시트는 Excel에서 한 번 열어 저장해야 마지막 계산값이 캐시됩니다."),
    ("참조 시트", "_Ref_ 시트의 이름을 바꾸거나 삭제하면 연결된 데이터 유효성 드롭다운이 깨질 수 있습니다."),
    ("CSV", "CSV는 자동 생성물입니다. 직접 수정하지 말고 이 워크북을 편집하세요."),
)


def decode_csv(raw: bytes, path: Path) -> str:
    """BOM을 우선해 혼재된 UTF-16/UTF-8 CSV를 판별한다."""
    if raw.startswith((b"\xff\xfe", b"\xfe\xff")):
        return raw.decode("utf-16")
    if raw.startswith(b"\xef\xbb\xbf"):
        return raw.decode("utf-8-sig")
    try:
        return raw.decode("utf-8")
    except UnicodeDecodeError as exc:
        raise UnicodeError(f"지원하지 않는 CSV 인코딩입니다(BOM 확인 필요): {path}") from exc


def read_csv(path: Path) -> list[list[str]]:
    text = decode_csv(path.read_bytes(), path)
    return [list(row) for row in csv.reader(io.StringIO(text, newline=""))]


def style_data_sheet(worksheet) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    fill = PatternFill("solid", fgColor="203864")
    for cell in worksheet[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    for column_index, cells in enumerate(worksheet.iter_cols(), start=1):
        width = max((len(str(cell.value)) if cell.value is not None else 0) for cell in cells)
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(max(width + 2, 10), 80)


def add_readme(workbook: Workbook) -> None:
    worksheet = workbook.create_sheet("_README", 0)
    for row in README_ROWS:
        worksheet.append(row)
    worksheet.column_dimensions["A"].width = 20
    worksheet.column_dimensions["B"].width = 110
    worksheet["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    worksheet["B1"].fill = worksheet["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    for row in worksheet.iter_rows(min_row=2):
        row[0].font = Font(bold=True)
        row[1].alignment = Alignment(wrap_text=True, vertical="top")


def create_workbook(
    excel_directory: Path,
    workbook_name: str,
    sheet_names: tuple[str, ...],
    csv_directory: Path | None = None,
) -> Path:
    workbook = Workbook()
    workbook.remove(workbook.active)
    add_readme(workbook)

    source_directory = csv_directory or excel_directory
    for sheet_name in sheet_names:
        csv_path = source_directory / f"{sheet_name}.csv"
        if not csv_path.is_file():
            raise FileNotFoundError(f"부트스트랩 원본 CSV를 찾을 수 없습니다: {csv_path}")
        rows = read_csv(csv_path)
        if not rows:
            raise ValueError(f"CSV가 비어 있습니다: {csv_path}")

        worksheet = workbook.create_sheet(sheet_name)
        for row in rows:
            worksheet.append(row)
        style_data_sheet(worksheet)

    output_path = excel_directory / workbook_name
    try:
        workbook.save(output_path)
    except PermissionError as exc:
        raise RuntimeError(f"엑셀 파일이 열려 있거나 잠겨 있습니다: {output_path}") from exc
    return output_path


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--excel-dir", type=Path, default=Path("LastFPS/Excel"))
    # 생성된 CSV는 워크북과 분리된 폴더에 있으므로 원본 위치를 따로 받는다.
    parser.add_argument("--csv-dir", type=Path, default=None)
    parser.add_argument("--force", action="store_true", help="기존 워크북을 덮어씁니다.")
    args = parser.parse_args()

    excel_directory = args.excel_dir.resolve()
    csv_directory = (args.csv_dir.resolve() if args.csv_dir else excel_directory / "Csv")
    for workbook_name, sheet_names in WORKBOOK_SHEETS.items():
        output_path = excel_directory / workbook_name
        if output_path.exists() and not args.force:
            raise FileExistsError(f"기존 워크북을 덮어쓰지 않습니다: {output_path} (--force 필요)")
        print(f"생성 완료: {create_workbook(excel_directory, workbook_name, sheet_names, csv_directory)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
