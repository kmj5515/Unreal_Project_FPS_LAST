#!/usr/bin/env python3
"""EditorUtility C++ 서비스가 호출하는 xlsx 파싱 진입점이다."""

from __future__ import annotations

import argparse
import importlib.util
import json
from pathlib import Path

from openpyxl import load_workbook


def load_converter(project_directory: Path):
    script_path = project_directory.parent / "Tools" / "Excel" / "xlsx_to_csv.py"
    if not script_path.is_file():
        raise FileNotFoundError(f"xlsx 변환 스크립트를 찾을 수 없습니다: {script_path}")
    spec = importlib.util.spec_from_file_location("lastfps_xlsx_to_csv", script_path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"xlsx 변환 스크립트를 로드할 수 없습니다: {script_path}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def scan_workbook(workbook_path: Path, output_path: Path) -> None:
    try:
        workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    except PermissionError as exc:
        raise RuntimeError(f"엑셀 파일이 열려 있거나 잠겨 있습니다: {workbook_path}") from exc
    try:
        payload = {"workbook": workbook_path.name, "sheets": workbook.sheetnames}
    finally:
        workbook.close()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--project-dir", type=Path, required=True)
    subparsers = parser.add_subparsers(dest="command", required=True)

    scan_parser = subparsers.add_parser("scan")
    scan_parser.add_argument("--workbook", type=Path, required=True)
    scan_parser.add_argument("--output", type=Path, required=True)

    convert_parser = subparsers.add_parser("convert")
    convert_parser.add_argument("--workbook", type=Path, required=True)
    convert_parser.add_argument("--sheet", required=True)
    convert_parser.add_argument("--output", type=Path, required=True)
    convert_parser.add_argument("--clean-output", type=Path, required=True)
    convert_parser.add_argument("--localize", action="store_true")

    args = parser.parse_args()
    if args.command == "scan":
        scan_workbook(args.workbook.resolve(), args.output.resolve())
    else:
        converter = load_converter(args.project_dir.resolve())
        # 언리얼 CSV 임포터에는 주석 문법이 없어 안내 행도 데이터 행으로 읽힌다.
        # 커밋용 CSV는 에셋의 리임포트 소스로도 쓰이므로, 안내 행을 넣으면
        # 우클릭 Reimport가 '#자동 생성됨...' RowName을 가진 행을 만들어낸다.
        # 편집 금지 안내는 각 워크북의 _README 시트가 담당한다.
        rows = converter.convert_sheet(
            args.workbook.resolve(),
            args.sheet,
            args.output.resolve(),
            localize=args.localize,
            include_notice=False,
        )
        converter.write_csv(rows, args.clean_output.resolve(), include_notice=False)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
